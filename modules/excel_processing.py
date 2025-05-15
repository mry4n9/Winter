import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from typing import List, Dict, Any

def create_excel_report(ad_data_dict: Dict[str, List[Dict[str, Any]]]) -> BytesIO:
    """
    Creates an XLSX report from ad data and returns it as a BytesIO stream.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet if it exists
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1: # Keep if it's the only sheet
        default_sheet = wb["Sheet"]
        wb.remove(default_sheet)
    elif "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1 and not ad_data_dict: # No data, keep default
        pass
    elif "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1 and ad_data_dict: # Data exists, remove default
         default_sheet = wb["Sheet"]
         wb.remove(default_sheet)


    header_font = Font(color="FFFFFF", bold=True) # White font
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Black fill
    header_alignment = Alignment(horizontal="center", vertical="center")
    content_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border_side = Side(style='thin', color="000000")
    cell_border = Border(left=thin_border_side,
                         right=thin_border_side,
                         top=thin_border_side,
                         bottom=thin_border_side)

    def style_header_cell(cell):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border
    
    def style_content_cell(cell):
        cell.alignment = content_alignment
        cell.border = cell_border

    # Email Page
    if 'Email' in ad_data_dict and ad_data_dict['Email']:
        ws_email = wb.create_sheet("Email")
        email_headers = ["Ad Name", "Funnel Stage", "Headline", "Subject Line", "Body", "CTA"]
        for col_num, header_text in enumerate(email_headers, 1):
            cell = ws_email.cell(row=1, column=col_num, value=header_text)
            style_header_cell(cell)
        
        for row_idx, ad in enumerate(ad_data_dict['Email'], 2):
            ws_email.cell(row=row_idx, column=1, value=ad.get("Ad Name", "")).border = cell_border
            ws_email.cell(row=row_idx, column=2, value=ad.get("Funnel Stage", "")).border = cell_border
            ws_email.cell(row=row_idx, column=3, value=ad.get("Headline", "")).border = cell_border
            ws_email.cell(row=row_idx, column=4, value=ad.get("Subject Line", "")).border = cell_border
            ws_email.cell(row=row_idx, column=5, value=ad.get("Body", "")).border = cell_border
            ws_email.cell(row=row_idx, column=6, value=ad.get("CTA", "")).border = cell_border
            for col_idx in range(1, len(email_headers) + 1):
                 style_content_cell(ws_email.cell(row=row_idx, column=col_idx))


    # LinkedIn Page
    if 'LinkedIn' in ad_data_dict and ad_data_dict['LinkedIn']:
        ws_linkedin = wb.create_sheet("LinkedIn")
        linkedin_headers = ["Ad Name", "Funnel Stage", "Introductory Text", "Image Copy", "Headline", "Destination", "CTA Button"]
        for col_num, header_text in enumerate(linkedin_headers, 1):
            cell = ws_linkedin.cell(row=1, column=col_num, value=header_text)
            style_header_cell(cell)

        for row_idx, ad in enumerate(ad_data_dict['LinkedIn'], 2):
            ws_linkedin.cell(row=row_idx, column=1, value=ad.get("Ad Name", "")).border = cell_border
            ws_linkedin.cell(row=row_idx, column=2, value=ad.get("Funnel Stage", "")).border = cell_border
            ws_linkedin.cell(row=row_idx, column=3, value=ad.get("Introductory Text", "")).border = cell_border
            ws_linkedin.cell(row=row_idx, column=4, value=ad.get("Image Copy", "")).border = cell_border
            ws_linkedin.cell(row=row_idx, column=5, value=ad.get("Headline", "")).border = cell_border
            ws_linkedin.cell(row=row_idx, column=6, value=ad.get("Destination", "")).border = cell_border
            ws_linkedin.cell(row=row_idx, column=7, value=ad.get("CTA Button", "")).border = cell_border
            for col_idx in range(1, len(linkedin_headers) + 1):
                 style_content_cell(ws_linkedin.cell(row=row_idx, column=col_idx))

    # FaceBook Page
    if 'Facebook' in ad_data_dict and ad_data_dict['Facebook']:
        ws_facebook = wb.create_sheet("Facebook")
        facebook_headers = ["Ad Name", "Funnel Stage", "Primary Text", "Image Copy", "Headline", "Link Description", "Destination", "CTA Button"]
        for col_num, header_text in enumerate(facebook_headers, 1):
            cell = ws_facebook.cell(row=1, column=col_num, value=header_text)
            style_header_cell(cell)

        for row_idx, ad in enumerate(ad_data_dict['Facebook'], 2):
            ws_facebook.cell(row=row_idx, column=1, value=ad.get("Ad Name", "")).border = cell_border
            ws_facebook.cell(row=row_idx, column=2, value=ad.get("Funnel Stage", "")).border = cell_border
            ws_facebook.cell(row=row_idx, column=3, value=ad.get("Primary Text", "")).border = cell_border
            ws_facebook.cell(row=row_idx, column=4, value=ad.get("Image Copy", "")).border = cell_border
            ws_facebook.cell(row=row_idx, column=5, value=ad.get("Headline", "")).border = cell_border
            ws_facebook.cell(row=row_idx, column=6, value=ad.get("Link Description", "")).border = cell_border
            ws_facebook.cell(row=row_idx, column=7, value=ad.get("Destination", "")).border = cell_border
            ws_facebook.cell(row=row_idx, column=8, value=ad.get("CTA Button", "")).border = cell_border
            for col_idx in range(1, len(facebook_headers) + 1):
                 style_content_cell(ws_facebook.cell(row=row_idx, column=col_idx))

    # Google Search Page
    if 'Google Search' in ad_data_dict and ad_data_dict['Google Search']:
        ws_gsearch = wb.create_sheet("Google Search")
        gsearch_headers = ["Headline", "Description"]
        for col_num, header_text in enumerate(gsearch_headers, 1):
            cell = ws_gsearch.cell(row=1, column=col_num, value=header_text)
            style_header_cell(cell)
        
        for row_idx, ad in enumerate(ad_data_dict['Google Search'], 2): # Max 15 ads
            ws_gsearch.cell(row=row_idx, column=1, value=ad.get("Headline", "")).border = cell_border
            ws_gsearch.cell(row=row_idx, column=2, value=ad.get("Description", "")).border = cell_border # Blank for rows 5-15
            for col_idx in range(1, len(gsearch_headers) + 1):
                 style_content_cell(ws_gsearch.cell(row=row_idx, column=col_idx))


    # Google Display Page
    if 'Google Display' in ad_data_dict and ad_data_dict['Google Display']:
        ws_gdisplay = wb.create_sheet("Google Display")
        gdisplay_headers = ["Headline", "Description"]
        for col_num, header_text in enumerate(gdisplay_headers, 1):
            cell = ws_gdisplay.cell(row=1, column=col_num, value=header_text)
            style_header_cell(cell)

        for row_idx, ad in enumerate(ad_data_dict['Google Display'], 2): # Max 5 ads
            ws_gdisplay.cell(row=row_idx, column=1, value=ad.get("Headline", "")).border = cell_border
            ws_gdisplay.cell(row=row_idx, column=2, value=ad.get("Description", "")).border = cell_border
            for col_idx in range(1, len(gdisplay_headers) + 1):
                 style_content_cell(ws_gdisplay.cell(row=row_idx, column=col_idx))
    
    # Auto-adjust column widths for all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter 
            for cell in col:
                try:
                    if cell.value:
                        cell_text_length = len(str(cell.value))
                        # For wrapped text, consider line breaks or estimate based on a max line width
                        # This is a simple approach; more sophisticated needed for perfect auto-fit with wrap
                        lines = str(cell.value).count('\n') + 1
                        max_line_length = 0
                        if lines > 1:
                            max_line_length = max(len(s) for s in str(cell.value).split('\n'))
                        else:
                            max_line_length = cell_text_length
                        
                        if max_line_length > max_length:
                            max_length = max_line_length
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2 # Add padding and factor
            if adjusted_width > 50: # Max width to prevent extremely wide columns
                adjusted_width = 50
            ws.column_dimensions[column_letter].width = adjusted_width


    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream